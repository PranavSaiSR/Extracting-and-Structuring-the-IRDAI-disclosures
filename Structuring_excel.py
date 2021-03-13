# -*- coding: utf-8 -*-
"""
Created on Thu Jul 16 15:56:41 2020

@author: PRANAVSAI
"""

# -*- coding: utf-8 -*-
"""
Created on Sun May 31 13:24:20 2020

@author: PRANAVSAI
"""
from openpyxl import Workbook, load_workbook
import numpy as np
import getpass


import os
dirListing = os.listdir(r'C:\data of comp\Sent already\1 Apollo Munic\excel\exc')
excel = []
for item in dirListing:
    if ".xlsx" in item:
        excel.append(item)

L = ["Form L-1-A-RA","Form L-2-A-PL","Form L-3- A-BS","Form L-4 Premium Schedule",
     "Form L-5- Commission Schedule","L-6- Operating Expenses Schedule",
     "Form L-7- Benefits Paid Schedule",
     "Form L-8 - Share Capital Schedule",
     "Form L-9- Pattern of Shareholding Schedule",
     "Form L-10-Reserves and Surplus Schedule",
     "Form L-11 -Borrowings Schedule",
     "Form L-12- Investments- Shareholders Schedule",
     "Form L-13-Investments- Policyholder Schedule",
     "Form L-14- Assets held to cover Linked Liabilities Schedule",
     "Form L-15-Loans Schedule",
     #"Form L-16-Property Plant & Equipments",
     "Form L-16- Fixed Assets Schedule",
     "Form L-17-Cash and Bank Balances Schedule",
     "Form L-18-Advances and Other Assets Schedule",
     "Form L-19-Current Liabilities Schedule",
     "Form L-20- Provisions Schedule",
     "Form L-21-Miscellaneous Expenditure Schedule",
     #"Form L-21-Misc Expenditure Schedule",
     "FORM L-22 : Analytical Ratios",
     "FORM L-23 RECEIPT AND PAYMENTS SCHEDULE",
     "FORM L-24  Valuation of net Liabilities",
     "FORM L-25- (i)- Geographical Distribution Channel - Individuals",
     "FORM L-26-INVESTMENT ASSETS(LIFE INSURERS)-3A",
     "FORM L-27-UNIT LINKED BUSINESS-3A",
     "FORM L-28-ULIP-NAV-3A",
     "FORM L-29 - Detail regarding debt securities (Non Ulip)",
     "FORM L-30 : Related Party Transactions",
     "FORM - L-31",
     "FORM L-32-SOLVENCY MARGIN - KT 3",
     "FORM L-34-YIELD ON INVESTMENTS",
     "FORM L-35-DOWNGRADING OF INVESTMENTS",
     "FORM L-36",
     "FORM L-37-BUSINESS ACQUISITION THROUGH DFFERENT CHANNELS (GROUP)",
     "FORM L-38- BUSINESS ACQUISITION THROUGH DIFFERENT CHANNELS (INDIVIDUALS)",
     "FORM L-39-Data on Settlement of Claims",
     "FORM L-40  : Quarterly claims data for Life",
     "FORM L-41 - GREIVANCE DISPOSAL",
     "L-42- Valuation Basis (Life Insurance)"]


NL = ["FORM NL-2-B-PL","FORM NL-3-B-BS","FORM NL-4-PREMIUM SCHEDULE",
      "FORM NL-10-RESERVE AND SURPLUS SCHEDULE",
#      "FORM NL-10-RESERVE AND SURPLUS SCHEDULE",
#      "FORM NL-12-INVESTMENT SCHEDULE",
#      "FORM NL-12 (B)-INVESTMENT SCHEDULE (SHAREHOLDERS)",
#      "From NL 12 : Investments - Shareholders",
      "Reinsurance Risk Concentration",
#      "FORM NL-36-YIELD ON INVESTMENTS 1",
      "FORM NL-36-YIELD ON INVESTMENTS 1"]



NL = ["OPERATING PROFIT/(LOSS)","SOURCES OF FUNDS","FORM NL-4-PREMIUM SCHEDULE",
      "FORM NL-10-RESERVE AND SURPLUS SCHEDULE",
#      "FORM NL-12-INVESTMENTS SCHEDULE",
#      "FORM NL-12-INVESTMENT SCHEDULE- SHAREHOLDERS",
#      "REINSURANCE RISK CONCENTRATION",
      "Reinsurance Risk Concentration",

#      "FORM NL-23",
      "Statement of Investment and Income on Investment"]


def wordfinder2(searchString,row,ind):

    for i in range(row, ws.max_row + 1):
        for j in range(1, 30):
            tmp = damerau_levenshtein_distance(searchString,str(ws.cell(i,j).value))
            
            if searchString == "FORM NL-23":
                if tmp <1:
                    print("found")
                    print(ws.cell(i,j).value)
                    
                    print(i,j)
                    return(i)                
                
            
            
            else:
                if ind < 2:
                    
                    if tmp <3:
                        print("found")
                        print(ws.cell(i,j).value)
                        
                        print(i,j)
                        return(i)
                else:
                    if tmp <5:
                        print("found")
                        print(ws.cell(i,j).value)
                        
                        print(i,j)
                        return(i) 



 
#File to be pasted into
Tmp = Workbook() #Add file name
for i in range(42):
#    if i != 25 or i!= 31 or i!=33:
    wb.create_sheet(str(1))
        

wb.save("D:\Life Data\Edelweiss Tokio Life Insurance Co. Ltd\PDFs\Pd-14aug_14-8-2019-2019-10-3--14-23-39-843-2019-11-22--17-3-15-495.xlsx")                 

def damerau_levenshtein_distance(s1, s2):
    d = {}
    lenstr1 = len(s1)
    lenstr2 = len(s2)
    for i in range(-1,lenstr1+1):
        d[(i,-1)] = i+1
    for j in range(-1,lenstr2+1):
        d[(-1,j)] = j+1

    for i in range(lenstr1):
        for j in range(lenstr2):
            if s1[i] == s2[j]:
                cost = 0
            else:
                cost = 1
            d[(i,j)] = min(
                           d[(i-1,j)] + 1, # deletion
                           d[(i,j-1)] + 1, # insertion
                           d[(i-1,j-1)] + cost, # substitution
                          )
            if i and j and s1[i]==s2[j-1] and s1[i-1] == s2[j]:
                d[(i,j)] = min (d[(i,j)], d[i-2,j-2] + cost) # transposition

    return d[lenstr1-1,lenstr2-1]

####################################################
index=0
for index1 in range(len(excel)):
    exc=excel[index]
    wb = load_workbook("C:\data of comp\Sent already\\1 Apollo Munic\excel\exc\\"+exc)
    ws = wb.get_sheet_by_name("Sheet1")  
    print(f'index={index}')
    
    a= []
    for i in range(len(NL)):
        if i == 0:   
            tmp=wordfinder2(NL[i],1,i)
            print("----")
            a.append(tmp)
            tmp1=a[i]
        else:
            tmp=wordfinder2(NL[i],tmp1,i)
            print(NL[i])
            print("----")
            a.append(tmp)
            tmp1=a[i]
        print(i)
    a.append(ws.max_row)
    for i in range(42):
        #    if i != 25 or i!= 31 or i!=33:
        wb.create_sheet(str(i+1))

    
    k=1
    l=1
    for ind in range(len(a)-1):
        k=1   
        
        if ind< 32:
            
            temp=a[ind]
            ts= wb.get_sheet_by_name(str(ind+1))
            for i in range(a[ind], a[ind+1]-1):
                l=1
                for j in range(1,75):
                    ts.cell(k,l).value = ws.cell(i, j).value            
                    l=l+1
                k=k+1
        else:
            temp=a[ind]
            ts= wb.get_sheet_by_name(str(ind+2))
            for i in range(a[ind], a[ind+1]-1):
                l=1
                for j in range(1,30):
                    ts.cell(k,l).value = ws.cell(i, j).value            
                    l=l+1
                k=k+1
        print(f'sheet={ind}')
    
    
    wb.save("C:\data of comp\Sent already\\1 Apollo Munic\excel\exc\\" + exc)    
    print(f'{index} = index')             
    index=index+1

###################################################
def wordfinder(searchString,row,maxr):
    tmp=[]
    rmp=[]
    cmp=[]
        
    
    for i in range(row, maxr + 1):
        for j in range(1, ls.max_column + 1):
            tmp1 = jaro_Winkler(searchString,str(ls.cell(i,j).value))
            tmp.append(tmp1)
            rmp.append(i)
            cmp.append(j)
            if searchString != "Sub-Total (B)":
                if tmp1 >= .9:
                    print("found")
                    print(ls.cell(i,j).value)
                    
                    print(i,j)
                    return(i,j)
                    
            if searchString == "Sub-Total (B)":
                if tmp1 == 1:
                    print("found")
                    print(ls.cell(i,j).value)
                    
                    print(i,j)
                    return(i,j)
            
    tmp2= tmp.index(max(tmp))
    rind = rmp[tmp2]
    cind = cmp[tmp2]
    return rind,cind


def colfinder(exc,i):
#    if "q1" in exc:
#        tmp = "30th June"
#    if "q2" in exc:
#        tmp = "30th Sep"    
#    if "q3" in exc:
#        tmp = "31st Dec"    
#    if "q4" in exc:
#        tmp = "31st Mar"
        
    if "q1" in exc:
        tmp = "June 30 ,"
    if "q2" in exc:
        tmp = "September 30 ,"    
    if "q3" in exc:
        tmp = "December 31 ,"    
    if "q4" in exc:
        tmp = "March 31 ,"
       
    if i == 1 or i ==3  :
        tmp = tmp +" 20"+ exc[:2]
#    if i==3:
#        tmp= "For The Quarter Ended" +tmp+ " 20"+ exc[:2]
    if i==4 or i==5 or i==2:
        tmp= "As at " +tmp+ " 20"+ exc[:2]
    r1,c1= wordfinder(tmp,1,10) 
    return r1,c1


def indfinder(searchString,row):

    i=row
    tmp=[]
    cmp=[]

    for j in range(1, fs.max_column + 1):
        word=searchString[:5]
        tmp1 = jaro_Winkler(word,str(fs.cell(i,j).value))
        tmp.append(tmp1)
        cmp.append(j)
    
    tmp2= tmp.index(max(tmp))
    cind = cmp[tmp2]

    return 0,cind
    
        
                
def colind(exc,i):

        
    r1,c1= indfinder(exc,1)
    return c1
#lb = load_workbook("C:\data of comp\Sent already\Aditya Birla\17 q3.xlsx\\")

lb = load_workbook("C:\data of comp\Sent already\Aditya Birla\exc\\"+exc)

ls= lb.get_sheet_by_name(str(1))        
row,dum=wordfinder("Profit / (Loss) Before Tax",1,ls.max_row) 
dum2,col= colfinder(exc,1)

ls.cell(row,col).value

pos= col +2
r=row
sam=[]  

for i in range(38):
    sam.append(ls.cell(r,pos).value)
    r=r+1
    
ls= Tmp.get_sheet_by_name(str(6))        
row,col=wordfinder("Employees' remuneration & welfare benefits",1) 

pos= col +1
r=row
sam=[]  

for i in range(30):
    sam.append(ls.cell(r,pos).value)
    r=r+1




    
k=1
l=2
for i in range(15,21):
    for j in range(4):
        
        tmp = str(i) + " q"+str(j+1)
        fs.cell(k,l).value=tmp
        l=l+1

fb.save("C:\data of comp\Sent already\Aditya Birla\exc\Final Birla.xlsx")
################################
################################
################################
################################
fb = load_workbook("C:\data of comp\Sent already\\2 ECGC\ECGC\merged\pdfs\exc\\\Final ECGC.xlsx")

fs = fb.get_sheet_by_name("Sheet1")




index=0
a=[]
for ind in range(len(excel)):
    exc = excel[index]
    
    lb = load_workbook("C:\data of comp\Sent already\\2 ECGC\ECGC\merged\pdfs\exc\\"+exc)
    ls= lb.get_sheet_by_name(str(1))  
 
#    d1,col= colfinder(exc,1)
#    row=row+1
    row,dum=wordfinder("Profit Before Tax ( A - B)",1,ls.max_row) 

    if ls.cell(row,dum+1).value is None:        
        col=dum+2
    else:
        col=dum+1
    if ls.cell(row,dum+2).value is None:
        col= dum+3

#    col=2
    tmp= ls.cell(row,col).value
    a.append(tmp)
    rpos=2
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
#    index=index+1
#    
#    
#index=0
#a=[]
#for ind in range(len(excel)):
#    exc = excel[index]
#    
#    lb = load_workbook("C:\data of comp\Sent already\Manipal Cigna 0\exc\\"+exc)
    ls= lb.get_sheet_by_name(str(2))        
   
    row,dum=wordfinder("Share Capital",1,ls.max_row)
#    row,dum=wordfinder("SHARE CAPITAL",1,ls.max_row)
    dum1,col=wordfinder("Schedule",1,ls.max_row)
    row=row+1
    col=col+1
    if ls.cell(row,col).value is None:
        col=col+1
#    if ls.cell(row,dum+1).value is None:
#        col=dum+3
#    else:
#        col=dum+2
            
        
    
#    row=row+1
#    col=3
    
    tmp= ls.cell(row,col).value
    a.append(tmp)
    rpos=3
    cpos=colind(exc,2)
    
    fs.cell(rpos,cpos).value = tmp
#    index=index+1
    
    
    row,dum=wordfinder("Net Current Assets ( C )= (A-B)",1,ls.max_row)
    row=row-1
#    dum1,col=wordfinder("NL-8-Share Capital Schedule",1,ls.max_row)

#    row = row
#    d1,col= colfinder(exc,2)
#    if ls.cell(row,col+1).value is None:
#        col=col+2
#    else:
#        col=col+1
    tmp= ls.cell(row,col).value
    if tmp is None:
        tmp= ls.cell(row,col+1).value
    a.append(tmp)
    rpos=4
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
    
    
    row,dum=wordfinder("Net Current Assets ( C )= (A-B)",1,ls.max_row) 
#    d1,col= colfinder(exc,2)
#    dum1,col=wordfinder("NL-8-Share Capital Schedule",1,ls.max_row)
#
#    if ls.cell(row,col+1).value is None:
#        col=col+2
#    else:
#        col=col+1
    tmp= ls.cell(row,col).value
    if tmp is None:
        tmp= ls.cell(row,col+1).value
    a.append(tmp)
    rpos=5
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
    
    row,dum=wordfinder("Total",1,ls.max_row) 
#    row,dum=wordfinder("TOTAL",row,ls.max_row)
    
    
#    dum1,col=wordfinder("NL-8-Share Capital Schedule",1,ls.max_row)
#
##    row=row+1
##    d1,col= colfinder(exc,2)
#    if ls.cell(row,col+1).value is None:
#        col=col+2
#    else:
#        col=col+1    
    tmp= ls.cell(row,col).value
    if tmp is None:
        tmp= ls.cell(row,col-1).value
    a.append(tmp)
    rpos=6
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
    
#    index=index+1
#
#index=0
#a=[]
#for ind in range(len(excel)):
#    exc = excel[1]
    
#    lb = load_workbook("C:\data of comp\Sent already\Aditya Birla\exc\\"+exc)
    ls= lb.get_sheet_by_name(str(3))   
    if index<100:     
        row,dum=wordfinder("Premium from Direct Business  Written",1,ls.max_row)
        if ls.cell(row,dum+1).value is None:        
            col=dum+2
        else:
            col=dum+1
        if ls.cell(row,dum+2).value is None and ls.cell(row,dum+1).value is None:
            col= dum+3
#        
    else:
        row,dum1=wordfinder("Premium from direct",1,ls.max_row)
        dum,col=wordfinder("Total",1,ls.max_row)
        if ls.cell(row,col).value is None:
            row=row+1
    
#    if index<10:
#    row,dum=wordfinder("Premium from direct business written",dum1,ls.max_row)
#    dum1,col=wordfinder("Total",dum1,ls.max_row)

#    col=dum+1
#    else:
#
#        dum1,col=wordfinder("Total",1,ls.max_row)   
#
#        row,dum=wordfinder("Premium from direct business written",dum1,ls.max_row)
    



#    if ls.cell(row,dum+1).value is None:        
#        col=dum+2
#    else:
#        col=dum+1
#    if ls.cell(row,dum+2).value is None and ls.cell(row,dum+1).value is None:
#        col= dum+3        


    
#    col=dum1+1
    tmp= ls.cell(row,col).value
    if tmp is None:
        col=col+1
        tmp= ls.cell(row,col).value

#    if tmp is None:
#        row=row-1
#        tmp= ls.cell(row,col).value
     
    a.append(tmp)
    rpos=7
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
#    index=index+1  
#index=0
#    
#a=[]
#for ind in range(len(excel)):
#    exc = excel[index]
    
#    lb = load_workbook("C:\data of comp\Sent already\Aditya Birla\exc\\"+exc)
    ls= lb.get_sheet_by_name(str(4))        
    row,dum=wordfinder("TOTAL",1,ls.max_row) 
#    d1,col= colfinder(exc,4)
    if ls.cell(row,dum+1).value is None:        
        col=dum+2
    else:
        col=dum+1
    if ls.cell(row,dum+2).value is None and ls.cell(row,dum+1).value is None:
        col= dum+3

    tmp= ls.cell(row,col).value
#    if tmp is None:
#        col=col-1
#        tmp= ls.cell(row,col).value
        
    a.append(tmp)
    rpos=8
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
#    index=index+1
#
#



#fb = load_workbook("C:\data of comp\Sent already\\1 Apollo Munic\excel\exc\Final Apollo.xlsx")
#
#fs = fb.get_sheet_by_name("Sheet1")
#
#
#
#index=0
#    
#for ind in range(len(excel)):
#    exc = excel[index]
#    
#    lb = load_workbook("C:\data of comp\Sent already\\1 Apollo Munic\excel\exc\\"+exc)
    ls= lb.get_sheet_by_name(str(6))        
##    row,dum=wordfinder("Total",1,ls.max_row)
    dum,col=wordfinder("Net Yield",1,ls.max_row)
    col=5
    row,dum1=wordfinder("GRAND TOTAL",1,ls.max_row) 
    col=col-1
        
#    if tmp == '-':
#        row,dum1=wordfinder("Total",1,ls.max_row)
    tmp= ls.cell(row,col).value
#    if tmp is None:
#        tmp= ls.cell(row,col-1).value
        
  
#    if ls.cell(row,col+1).value is None:
#        col=col+2
#    else:
#        col=col+1


    a.append(tmp)
    rpos=9
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
#    index=index+1
#fb.save("C:\data of comp\Sent already\\1 Apollo Munic\excel\exc\Final Apollo.xlsx")



#index=0
#a=[]   
#for ind in range(len(excel)):
#    exc = excel[index]
#    
#    lb = load_workbook("C:\data of comp\Sent already\Bajaj Allianz\pdfs\exc\\"+exc)
    ls= lb.get_sheet_by_name(str(5))        
    row1,dum1=wordfinder("Proportional",1,ls.max_row)
#    
###    row2,dum2=wordfinder("Non-Proportional",1,ls.max_row)
    dum1=dum1
    
    row2=row1
    dum2=dum1+1
    
    row3=row1
    dum3=dum1+2
    row,dum=wordfinder("Total",row1+1,ls.max_row)
    
    
#    if row1==row2 and row1==row3:
    
            
        

#    if ls.cell(row,dum+1).value is None:
#        dum=dum+1
#        else:
#            row,dum=wordfinder("Grand Total (C)= (A)+(B)",row1+1,ls.max_row)
     
    
#        tmp1= ls.cell(row,dum1).value
#        if tmp1=="Total":
#            dum1=dum1+1
#            dum2=dum2+1
#            dum3=dum3+1
#            tmp1= ls.cell(row,dum1).value

#    dum1=dum+2
#    dum2=dum+4
#    dum3=dum+6
    tmp1= str(ls.cell(row,dum1).value)
#    if tmp1 == "Total":
#        tmp1=dum1
#        dum1=tmp1+2
#        dum2=tmp1+3
#        dum3=tmp1+4
#        tmp1= str(ls.cell(row,dum1).value)
    tmp2= str(ls.cell(row,dum2).value)
#
#    if tmp2 == "Total":
#        tmp1=dum2
#        dum1=tmp1+2
#        dum2=tmp1+3
#        dum3=tmp1+4
#        tmp1= str(ls.cell(row,dum1).value)
#        tmp2= str(ls.cell(row,dum2).value)
#
#    tmp3= str(ls.cell(row,dum3).value)
    tmp3=0
#
#
    if ',' in tmp1:
        tmp1=tmp1.replace(',', '')
    if ',' in tmp2:
        tmp2=tmp2.replace(',', '')
#    if ',' in tmp3:
#        tmp3=tmp3.replace(',', '')    
    if tmp1 is None or tmp1 == '-' or tmp1 == '#':
        tmp1 = 0
    if tmp2 is None or tmp2 == '-':
        tmp2=0          
    if tmp3 is None or tmp3 == '-':
        tmp3=0
#    tmp =float(tmp1)+float(tmp2)+float(tmp3)    
    tmp = ls.cell(row,dum1+3).value
    a.append(tmp)
    if str(tmp3) == '1':
        tmp="check"
#        
    rpos=10
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
#    index=index+1
##
##
#index=0
##a=[]
#for ind in range(len(excel)):
#    exc = excel[index]
#    
#    lb = load_workbook("C:\data of comp\Sent already\AXA\pdfs\merged\pdfs\\"+exc)
    ls= lb.get_sheet_by_name(str(6))     
    dum,col=wordfinder("Income",1,ls.max_row) 
    row,dum1=wordfinder("GRAND TOTAL",1,ls.max_row) 
    
##    col=ls.max_column -12
#    if ls.cell(row,col).value is None:
#        col=col+1
#    else:
#        col=col
    
#    row=row+1
#    col=col-1
    tmp= ls.cell(row,col).value
#    if tmp == '-':
#        row,dum1=wordfinder("Total",1,ls.max_row)
#        tmp= ls.cell(row,col).value

    a.append(tmp)
    rpos=11
    cpos=colind(exc,1)
    
    fs.cell(rpos,cpos).value = tmp
    index=index+1




fb.save("C:\data of comp\Sent already\\2 ECGC\ECGC\merged\pdfs\exc\\\Final ECGC.xlsx")
   
    

          